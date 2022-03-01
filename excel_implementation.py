from openpyxl import Workbook
from excel_abstraction import EmployeeAbstraction
from excel_classes import Address, Employee
from excelfile import get_wb_and_sheet,file_path
from openpyxl.styles import Alignment
from excel_emp_data import generate_emp


class empimplementation(EmployeeAbstraction):
    
    def writeheader(self):
        workbook,sheet = get_wb_and_sheet()
        header_line = "id name age salary address".split()
        for i in range(1,len(header_line)+1):
            sheet.cell(row=1,column=i).value=header_line[i-1]
        sheet.merge_cells('E1:G1')
        
        address_header_line = "pincode city state".split()
        for i in range(5,8):
            sheet.cell(row=2,column=i).value=address_header_line[i-5]
        
        for i in range(1,sheet.max_row+1):
            for j in range(1,sheet.max_column+1):
                cell = sheet.cell(row=i,column=j)
                cell.alignment = Alignment(horizontal="center",vertical="center")
        workbook.save(file_path)

        
    def write_data(self):
        workbook,sheet = get_wb_and_sheet()
        # print(workbook,sheet)
        employee_list = generate_emp(100)
        r = sheet.max_row + 1
        for emp in employee_list:
            sheet.cell(row=r, column=1).value = emp.Empid
            sheet.cell(row=r, column=2).value = emp.Empname
            sheet.cell(row=r, column=3).value = emp.Empage
            sheet.cell(row=r, column=4).value = emp.Empsalary
            sheet.cell(row=r, column=5).value = emp.Empaddress.pincode
            sheet.cell(row=r, column=6).value = emp.Empaddress.City
            sheet.cell(row=r, column=7).value = emp.Empaddress.state 
            r += 1
        workbook.save(file_path)


    def read_data(self):
        """ Read data from Excel file"""
        list_of_emps = []
        workbook, sheet = get_wb_and_sheet()
        for i in range(3, sheet.max_row+1):
            empid = sheet.cell(row=i, column=1).value
            empname = sheet.cell(row=i, column=2).value
            empage = sheet.cell(row=i, column=3).value
            empsalary = sheet.cell(row=i, column=4).value
            adrpin = sheet.cell(row=i, column=5).value
            adrcity = sheet.cell(row=i, column=6).value
            adrstate = sheet.cell(row=i, column=7).value
            emp = Employee(eid=empid, ename=empname, eage=empage, esalary=empsalary, address=Address(pin=adrpin, city=adrcity, state=adrstate))
            list_of_emps.append(emp)
        print(list_of_emps)

if __name__=="__main__":
    empl = empimplementation()
    # empl.writeheader()
    # empl.write_data()
    empl.read_data()













