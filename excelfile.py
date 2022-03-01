import openpyxl

file_path = r"C:\Users\ramgi\Documents\All_Python\Employee\emp_data.xlsx"

file_path = r"D:\Pc\git_practice\B6_G8\emp_data.xlsx"

def get_wb_and_sheet():

    # wb=openpyxl.Workbook()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    # sheet=wb.create_sheet("employee_data")
    # del wb["Sheet"]
    # sheet=wb.active
    # print(sheet)
    wb.save(file_path)
    return wb,sheet









